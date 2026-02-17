import openpyxl
import pandas as pd
from collections import defaultdict
import sys

def analyze_backtest(excel_path):
    """分析回測結果"""
    print("=" * 70)
    print("選擇權賣方策略回測報告分析")
    print("=" * 70)
    print()

    # 創建一個工作簿對象
    wb = openpyxl.load_workbook(excel_path)

    # 讀取 Tree View sheet
    tree_sheet = wb["Tree View"]

    # 收集各個月的數據
    monthly_data = defaultdict(lambda: {
        "total_pnl": 0,
        "positions": [],
        "combo_pnls": []
    })

    # 解析 Tree View 數據
    for row_idx, row in enumerate(tree_sheet.iter_rows(min_row=2, values_only=True), start=2):
        month = str(row[0]) if row[0] else ""
        position_type = str(row[1]) if row[1] else ""
        closing_price = row[8] if row[8] is not None else 0
        total_pnl = row[9] if row[9] is not None else 0
        combo_pnl = row[10] if row[10] is not None else 0

        # 這是月份總結行
        if position_type != "賣出買權" and position_type != "賣出賣權" and position_type != "買權避險單" and position_type != "賣權避險單":
            monthly_data[month]["total_pnl"] = total_pnl

        # 這是賣出買權行（包含組合部位損益）
        elif position_type == "賣出買權":
            monthly_data[month]["combo_pnls"].append(combo_pnl)

    # 統計分析
    monthly_pnls = [data["total_pnl"] for data in monthly_data.values()]
    total_months = len(monthly_pnls)

    if total_months == 0:
        print("沒有找到任何回測數據！")
        return

    # 計算基本統計
    total_pnl = sum(monthly_pnls)
    avg_monthly_pnl = total_pnl / total_months
    positive_months = sum(1 for pnl in monthly_pnls if pnl > 0)
    negative_months = sum(1 for pnl in monthly_pnls if pnl < 0)

    # 最大獲利和最大虧損
    max_profit = max(monthly_pnls)
    max_loss = min(monthly_pnls)

    # 標準差
    import statistics
    std_dev = statistics.stdev(monthly_pnls) if total_months > 1 else 0

    # 夏普比率（假設無風險利率為 0）
    sharpe_ratio = avg_monthly_pnl / std_dev if std_dev > 0 else 0

    # 最大回撤
    cumulative_pnl = 0
    max_drawdown = 0
    peak = 0
    for pnl in monthly_pnls:
        cumulative_pnl += pnl
        if cumulative_pnl > peak:
            peak = cumulative_pnl
        drawdown = peak - cumulative_pnl
        if drawdown > max_drawdown:
            max_drawdown = drawdown

    # 收益率的月度標準差
    returns = [pnl / 100000000 for pnl in monthly_pnls]  # 假設資本 1 億
    avg_return = sum(returns) / len(returns) if returns else 0
    return_std = statistics.stdev(returns) if len(returns) > 1 else 0

    print(f"【基本統計】")
    print(f"回測月數: {total_months} 個月")
    print(f"總損益: ${total_pnl:,.0f}")
    print(f"平均月損益: ${avg_monthly_pnl:,.0f}")
    print(f"最大單月獲利: ${max_profit:,.0f}")
    print(f"最大單月虧損: ${max_loss:,.0f}")
    print(f"月度標準差: ${std_dev:,.0f}")
    print()

    print(f"【表現分析】")
    print(f"獲利月份: {positive_months} 個 ({positive_months/total_months*100:.1f}%)")
    print(f"虧損月份: {negative_months} 個 ({negative_months/total_months*100:.1f}%)")
    print(f"獲勝率: {positive_months/total_months*100:.1f}%")
    print(f"夏普比率: {sharpe_ratio:.2f}")
    print(f"最大回撤: ${max_drawdown:,.0f}")
    print()

    # 計算年化回報率和年化波動率
    annual_return = avg_monthly_pnl * 12
    annual_volatility = std_dev * (12 ** 0.5)
    annual_sharpe = annual_return / annual_volatility if annual_volatility > 0 else 0

    print(f"【年化指標】")
    print(f"年化預期回報: ${annual_return:,.0f}")
    print(f"年化波動率: ${annual_volatility:,.0f}")
    print(f"年化夏普比率: {annual_sharpe:.2f}")
    print()

    # 月度明細
    print(f"【月度損益明細】")
    print("-" * 70)
    print(f"{'月份':<12} {'月度損益':<15} {'累積損益':<15} {'組合數目':<10}")
    print("-" * 70)

    cumulative = 0
    sorted_months = sorted(monthly_data.keys())
    for month in sorted_months:
        data = monthly_data[month]
        pnl = data["total_pnl"]
        cumulative += pnl
        combo_count = len(data["combo_pnls"])

        print(f"{month:<12} ${pnl:>10,.0f}   ${cumulative:>10,.0f}   {combo_count:<10}")

    print("-" * 70)
    print()

    # 組合部位損益分析
    all_combo_pnls = []
    for data in monthly_data.values():
        all_combo_pnls.extend(data["combo_pnls"])

    if all_combo_pnls:
        print(f"【組合部位損益分析】")
        print(f"總組合數: {len(all_combo_pnls)}")
        positive_combos = sum(1 for pnl in all_combo_pnls if pnl > 0)
        negative_combos = sum(1 for pnl in all_combo_pnls if pnl < 0)
        print(f"獲利組合: {positive_combos} 個 ({positive_combos/len(all_combo_pnls)*100:.1f}%)")
        print(f"虧損組合: {negative_combos} 個 ({negative_combos/len(all_combo_pnls)*100:.1f}%)")
        print(f"平均組合損益: ${sum(all_combo_pnls)/len(all_combo_pnls):,.0f}")
        print(f"最大組合獲利: ${max(all_combo_pnls):,.0f}")
        print(f"最大組合虧損: ${min(all_combo_pnls):,.0f}")
        print()

    # 讀取 Monthly P&L sheet
    pnl_sheet = wb["Monthly P&L"]
    monthly_values = []
    for row in pnl_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1] is not None:
            monthly_values.append(row[1])

    print("=" * 70)
    print("分析完成！")
    print("=" * 70)


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "backtest_results.xlsx"
    analyze_backtest(excel_path)
