from openpyxl import load_workbook

wb = load_workbook("test_fill.xlsx")
ws = wb.active

print("Colors in test_fill.xlsx:")
for row_num in range(1, 4):
    cell = ws.cell(row=row_num, column=1)
    print(f"Row {row_num}: {cell}")
    if cell.fill and hasattr(cell.fill, 'start_color'):
        print(f"  Fill start color: {cell.fill.start_color}")
        if hasattr(cell.fill.start_color, 'rgb'):
            print(f"  RGB: {cell.fill.start_color.rgb}")
    print()

wb.close()
