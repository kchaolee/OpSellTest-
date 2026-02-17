import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def export_to_excel(results: dict, output_path: str, config=None):
    wb = Workbook()
    wb.remove(wb.active)

    tree_sheet = wb.create_sheet("Tree View")

    headers = ["月份", "部位類型", "建立時間", "建立時加權指數", "賣出履約價", "買入履約價",
               "權利金點數", "權利金總額", "結算日收盤價", "總損益", "組合部位損益"]
    tree_sheet.append(headers)
    
    # 設定第 11 欄（組合部位損益）的寬度為 15
    tree_sheet.column_dimensions[get_column_letter(11)].width = 15
    
    for cell in tree_sheet[1]:
        cell.font = Font(bold=True)
    
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # 組合部位損益的樣式（黑色和紅色兩種）
    combo_font_black = Font(name="Verdana", size=11, color="000000")
    combo_font_red = Font(name="Verdana", size=11, color="FF0000")
    
    row_num = 2
    
    # 從 config 獲取權利金點數和契約乘數
    if config:
        get_sell_call_point = config.get("get_sell_call_point", 400)
        get_sell_put_point = config.get("get_sell_put_point", 600)
        cost_buy_call_point = config.get("cost_buy_call_point", 200)
        cost_buy_put_point = config.get("cost_buy_put_point", 200)
        multiplier = config.get("contract_multiplier", 50)
    else:
        get_sell_call_point = 400
        get_sell_put_point = 600
        cost_buy_call_point = 200
        cost_buy_put_point = 200
        multiplier = 50
    
    for key, month_result in results.items():
        if isinstance(key, tuple):
            year, month = key
            date_key = (year, month)
        else:
            year = None
            month = key
            date_key = month
            
        start_date = month_result.get("start_date")
        settlement = month_result.get("settlement_date")
        closing = month_result.get("closing_price", 0)
        
        month_label = f"{year}-{month:02d}" if year else f"{month:02d}"
        period = ""
        if start_date and settlement:
            period = f"{start_date.strftime('%m/%d')}~{settlement.strftime('%m/%d')}"

        # Month summary row
        tree_sheet.append([month_label, period, "", "", "", "", "", "", closing, month_result["total_pnl"], ""])
        row_num += 1

        for idx, pos in enumerate(month_result["positions"], 1):
            date_str = pos["date"].strftime("%Y-%m-%d")
            base_index = pos["base_index"]
            
            # 計算權利金總額
            call_prem = get_sell_call_point * multiplier
            put_prem = get_sell_put_point * multiplier
            call_hedge_cost = cost_buy_call_point * multiplier
            put_hedge_cost = cost_buy_put_point * multiplier
            
            # 計算該部位的4個單元損益
            call_pnl = pos.get("call_pnl", 0)
            put_pnl = pos.get("put_pnl", 0)
            call_spread_pnl = pos.get("call_spread_pnl", 0)
            put_spread_pnl = pos.get("put_spread_pnl", 0)
            
            # 計算組合部位總損益（該組合的4個部位損益加總）
            combo_pnl = call_pnl + put_pnl + call_spread_pnl + put_spread_pnl

            # 賣出買權
            row1 = [month_label, "賣出買權", date_str, base_index, pos["sell_call_strike"], "",
                    get_sell_call_point, call_prem, closing, call_pnl, combo_pnl]
            tree_sheet.append(row1)
            for col_idx in range(1, len(row1) + 1):
                tree_sheet.cell(row=row_num, column=col_idx).fill = light_green_fill
            # 設定組合部位損益欄位的樣式（使用同列的淡綠色背景）
            combo_cell = tree_sheet.cell(row=row_num, column=11)
            combo_cell.font = combo_font_black if combo_pnl > 0 else combo_font_red
            row_num += 1
            
            # 賣出賣權（不顯示組合部位損益）
            row2 = [month_label, "賣出賣權", date_str, base_index, pos["sell_put_strike"], "",
                    get_sell_put_point, put_prem, closing, put_pnl, ""]
            tree_sheet.append(row2)
            for col_idx in range(1, len(row2) + 1):
                tree_sheet.cell(row=row_num, column=col_idx).fill = light_green_fill
            row_num += 1
            
            # 買權避險單（不顯示組合部位損益）
            row3 = [month_label, "買權避險單", date_str, base_index, pos["call_buy_strike"], pos["call_sell_strike"],
                    cost_buy_call_point, call_hedge_cost, closing, call_spread_pnl, ""]
            tree_sheet.append(row3)
            for col_idx in range(1, len(row3) + 1):
                tree_sheet.cell(row=row_num, column=col_idx).fill = light_red_fill
            row_num += 1
            
            # 賣權避險單（不顯示組合部位損益）
            row4 = [month_label, "賣權避險單", date_str, base_index, pos["put_buy_strike"], pos["put_sell_strike"],
                    cost_buy_put_point, put_hedge_cost, closing, put_spread_pnl, ""]
            tree_sheet.append(row4)
            for col_idx in range(1, len(row4) + 1):
                tree_sheet.cell(row=row_num, column=col_idx).fill = light_red_fill
            row_num += 1

    pnl_sheet = wb.create_sheet("Monthly P&L")

    months = []
    pnls = []
    for key, month_result in results.items():
        if isinstance(key, tuple):
            year, month = key
            month_label = f"{year}-{month:02d}"
        else:
            month = key
            month_label = f"{month:02d}"
        months.append(month_label)
        pnls.append(month_result["total_pnl"])

    pnl_sheet.append(["月份", "總損益"])
    for month, pnl in zip(months, pnls):
        pnl_sheet.append([month, pnl])

    chart = BarChart()
    data = Reference(pnl_sheet, min_col=2, min_row=1, max_row=len(pnls)+1)
    cats = Reference(pnl_sheet, min_col=1, min_row=2, max_row=len(pnls)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "每月總損益"
    chart.x_axis.title = "月份"
    chart.y_axis.title = "損益 (元)"

    pnl_sheet.add_chart(chart, "D2")

    wb.save(output_path)
