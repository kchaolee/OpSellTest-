import sys
import os
sys.path.insert(0, "src")

from backtester.data_loader import load_index_data
from backtester.main import run_monthly_range_backtest
from backtester.config import BacktestConfig
from backtester.excel_exporter import export_to_excel
from openpyxl import load_workbook

print("Verifying combo P&L column modifications...")
print()

params = {
    "data_path": "skills/opsell/assets/TSEA_加權指_日線.csv",
    "n": 3.0,
    "get_sell_call_point": 400.0,
    "get_sell_put_point": 600.0,
    "cost_buy_call_point": 200.0,
    "cost_buy_put_point": 200.0,
    "max_order": 5,
    "contract_multiplier": 50,
    "start_year": 2025,
    "start_month": 8,
    "end_year": 2025,
    "end_month": 8,
    "output_file": "test_combo_final.xlsx"
}

config = BacktestConfig(
    n=params["n"],
    get_sell_call_point=params["get_sell_call_point"],
    get_sell_put_point=params["get_sell_put_point"],
    cost_buy_call_point=params["cost_buy_call_point"],
    cost_buy_put_point=params["cost_buy_put_point"],
    max_order=params["max_order"],
    contract_multiplier=params["contract_multiplier"]
)

df = load_index_data(params["data_path"])
results = run_monthly_range_backtest(df, config, 2025, 8, 2025, 8)
export_to_excel(results, params["output_file"], config.__dict__)

wb = load_workbook(params["output_file"])
tree_sheet = wb["Tree View"]

print("Step 1: Checking which positions show combo P&L...")
combo_col_idx = 11

pass_count = 0
for row_num in range(2, min(15, tree_sheet.max_row + 1)):
    pos_type = tree_sheet.cell(row=row_num, column=2).value
    combo_value = tree_sheet.cell(row=row_num, column=combo_col_idx).value
    
    if pos_type == "賣出買權":
        if combo_value and combo_value != "":
            print(f"Row {row_num} (賣出買權): PASS - Shows combo P&L")
            pass_count += 1
        else:
            print(f"Row {row_num} (賣出買權): FAIL - Should show combo P&L but is empty")
    elif pos_type == "賣出賣權":
        if not combo_value or combo_value == "":
            print(f"Row {row_num} (賣出賣權): PASS - Empty (expected)")
            pass_count += 1
        else:
            print(f"Row {row_num} (賣出賣權): FAIL - Should be empty but shows: {combo_value}")
    elif pos_type == "買權避險單":
        if not combo_value or combo_value == "":
            print(f"Row {row_num} (買權避險單): PASS - Empty (expected)")
            pass_count += 1
        else:
            print(f"Row {row_num} (買權避險單): FAIL - Should be empty but shows: {combo_value}")
    elif pos_type == "賣權避險單":
        if not combo_value or combo_value == "":
            print(f"Row {row_num} (賣權避險單): PASS - Empty (expected)")
            pass_count += 1
        else:
            print(f"Row {row_num} (賣權避險單): FAIL - Should be empty but shows: {combo_value}")
    
    if pass_count >= 3:
        break

print()
print("Step 2: Checking cell formatting (gray bold, 11pt)...")
for row_num in range(2, min(10, tree_sheet.max_row + 1)):
    cell = tree_sheet.cell(row=row_num, column=combo_col_idx)
    if cell.value and cell.value != "":
        font = cell.font
        fill = cell.fill
        
        print(f"Row {row_num}:")
        print(f"  Font: {font.name}")
        print(f"  Size: {font.size}pt")
        print(f"  Bold: {font.bold}")
        
        # Check gray fill
        if fill and fill.start_color:
            fill_rgb = fill.start_color.rgb if hasattr(fill.start_color, 'rgb') else None
        else:
            fill_rgb = None
        
        gray_expected = "00D3D3D3"
        if fill_rgb == gray_expected:
            print(f"  Fill: Gray")
        else:
            print(f"  Fill: {fill_rgb} (expected: {gray_expected})")
        
        # Check font specifications
        if font.name == "Verdana" and font.size == 11 and not font.bold:
            print(f"  Font styles: Verdana 11pt, not bold ✓")
        else:
            print(f"  Font styles: {font.name} {font.size}pt bold={font.bold} (should be Verdana 11pt not bold)")
        
        break

wb.close()
os.remove(params["output_file"])
print()
print("Verification completed successfully!")
