import os
import sys
sys.path.insert(0, "src")

import pytest
from backtester.data_loader import load_index_data
from backtester.main import run_yearly_backtest
from backtester.config import BacktestConfig
from backtester.excel_exporter import export_to_excel
import pandas as pd

def test_full_backtest_pipeline():
    data_path = "skills/opsell/assets/TSEA_加權指_日線.csv"
    output_path = "test_integration_results.xlsx"
    
    try:
        df = load_index_data(data_path)
        
        config = BacktestConfig(
            n=3.0,
            get_sell_call_point=400.0,
            get_sell_put_point=600.0,
            cost_buy_call_point=200.0,
            cost_buy_put_point=200.0,
            max_order=5,
        )
        
        results = run_yearly_backtest(df, config, 2023)
        
        assert results is not None
        assert len(results) == 12
        
        for month, result in results.items():
            assert "total_pnl" in result
            assert "positions" in result
            assert isinstance(result["positions"], list)
        
        export_to_excel(results, output_path)
        
        assert os.path.exists(output_path)
        
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        
        assert "Tree View" in wb.sheetnames
        assert "Monthly P&L" in wb.sheetnames
        
        tree_sheet = wb["Tree View"]
        headers = [cell.value for cell in tree_sheet[1]]
        expected_headers = ["月份", "部位類型", "建立時間", "建立時加權指數", "賣出履約價", "買入履約價",
                          "權利金點數", "權利金總額", "結算日收盤價", "總損益"]
        assert headers == expected_headers
        
        pnl_sheet = wb["Monthly P&L"]
        assert pnl_sheet["A1"].value == "月份"
        assert pnl_sheet["B1"].value == "總損益"
        
        wb.close()
        
    finally:
        if os.path.exists(output_path):
            os.remove(output_path)
