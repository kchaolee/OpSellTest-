import openpyxl
from openpyxl.styles import Font, PatternFill

def verify_excel():
    """Verify the Excel export meets all requirements"""
    wb = openpyxl.load_workbook("test_combo_final.xlsx")
    sheet = wb["Tree View"]

    print("Verification Results:")
    print("=" * 60)

    # Requirement 1 & 2: Only sold call shows combo P&L
    print("\n1. Check combo P&L column (column 11):")
    call_pnl = sheet.cell(row=3, column=11).value
    put_pnl = sheet.cell(row=4, column=11).value
    hedge1_pnl = sheet.cell(row=5, column=11).value

    print(f"   Row 3 (Call): {call_pnl if call_pnl else 'Empty'}")
    print(f"   Row 4 (Put): {put_pnl if put_pnl else 'Empty'}")
    print(f"   Row 5 (Hedge1): {hedge1_pnl if hedge1_pnl else 'Empty'}")

    if call_pnl and not put_pnl and not hedge1_pnl:
        print("   PASS: Only sold call shows combo P&L")
    else:
        print("   FAIL: Incorrect combo P&L display")

    # Requirement 3: Font is Verdana 11pt, not bold
    print("\n2. Check font formatting:")
    combo_cell = sheet.cell(row=3, column=11)
    font = combo_cell.font
    print(f"   Font name: {font.name}")
    print(f"   Font size: {font.size}")
    print(f"   Bold: {font.bold}")

    if font.name == "Verdana" and font.size == 11 and not font.bold:
        print("   PASS: Verdana 11pt, not bold")
    else:
        print("   FAIL: Incorrect font formatting")

    # Requirement 4: Background is gray
    print("\n3. Check cell background:")
    fill = combo_cell.fill
    print(f"   Fill color: {fill.start_color}")

    if fill.start_color and "D3D3D3" in str(fill.start_color):
        print("   PASS: Gray background")
    else:
        print("   FAIL: Incorrect background color")

    # Requirement 5: Column width is set to 15
    print("\n4. Check column width:")
    col_dim = sheet.column_dimensions['K']
    print(f"   Column K width: {col_dim.width}")

    if col_dim.width == 15:
        print("   PASS: Column width is 15")
    else:
        print("   FAIL: Incorrect column width")

    print("\n" + "=" * 60)
    print("Verification complete!")

if __name__ == "__main__":
    verify_excel()
